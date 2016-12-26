from __future__ import division		# All divisions are float point divisions by default. Useful in OU_line() when eval() UK odds.
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font				# for strikethroughing postponed games since openpyxl can't delete the row
from selenium import webdriver
from difflib import SequenceMatcher
from operator import itemgetter
import datetime
import re
import time


# ------------------------------------------ FUNCTIONS --------------------------------------------

# Returns soup form oddportal.com/"suffix" , given a "suffix"
def oddsportal_soup(suffix):

	user_agent = 'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/48.0.2564.116 Safari/537.36'

	webdriver.DesiredCapabilities.PHANTOMJS['phantomjs.page.customHeaders.User-Agent'] = user_agent 

	'''
	OddsPortal.com seems to block PhantomJS's browser (returns Page Not Found).
	Changing the header seems to succesfully do the trick and load the page's javascript driven content.
	This is what the above two lines do, change the User-Agent browser's header.
	'''

	driver = webdriver.PhantomJS()

	driver.get("http://www.oddsportal.com" + suffix)

	time.sleep(4)		# JavaScript takes a while to load sometimes. Wait for it.

	content = driver.page_source

	soup = BeautifulSoup(content,  "lxml")

	return soup


# Returns the soup of the main odportal.com soccer games page for given date
def games_page_for_date(date):

	return oddsportal_soup("/matches/soccer/" + date)


# Returns a dictionary of all the game URLs for given soup. Key: "teams' names" - Value: "game URL", e.g. dict["Tottenham Hotspur Swansea City"] = "/soccer/england/premier-league/tottenham-swansea-xWKLbeh8/"
def links_dict(soup):

	links = soup(attrs = {"class":"deactivate"})

	links_dict = {}

	for i in range (len(links)):

		temp = links[i].find("a", href = True)['href']		# Check if <a> has a "hfer" and if so, it grabs it. (clue: all hrefs are True here, just put it for completeness reasons)

		links_dict[temp.split("/")[4][:-9]] = temp

	return links_dict


# Returns similarity ratio between two strings. Called in find_game_in_page().
def similarity(a, b):
	    return SequenceMatcher(None, a, b).ratio()		# FYI, quick_ratio() and very_quick_ratio() available. Faster but compare less data - not suitable here.


# Returns the matching game URL, given links_dict and the game name as str().
def find_game_link_in_page(links_dict, game):

	ratios = {}

	for link in links_dict:

		ratios[link] = (similarity(game, link))

			
	'''# For testing purposes - Prints the whole dictionary sorted by ratio
				od = sorted(ratios.iteritems(), key = itemgetter(1))		# Used key=lambda x: x[1], but itemgetter() is faster.
				print(od[-2:])'''
	

	picked_match = max(ratios.iteritems(), key=itemgetter(1))[0]

	print(picked_match)

	if not any(word in picked_match for word in game.lower().split()):		# split - and . PLUS check words 3 letters or higher
		print("GAME DID NOT MATCH! Chack Date.")
		return

	picked_match_link = links_dict[picked_match]

	return picked_match_link

# Returns score as a tuple, given a soup and a game's href. Empty tuple if the game has been postponed.
def get_score(soup, link):

	try:
		find_score = soup.find(href=re.compile(link)).parent.parent.find_next('td', 'table-score').text.encode('utf-8')		# Gets href's parent tag and searches for the score within it.
	
	except AttributeError:																							# If there is no score posted yet
		print("Game not played yet.")
		return ()
	
	try:
		score_split = re.findall(r'\d+:\d+',find_score)[0].split(":")		# Accepts results like (1:3) (12:6) (0:2 pen.)
	
		score = (int(score_split[0]), int(score_split[1]))
		#score = (int(x) for x in score_split)

	except IndexError:														# For (award.) (postp.) (abn.) as result
		print("Game postponed. Marked for deletion.")
		return

	return score



# Returns a (OU_line, over_odds) tuple given a game link.
def OU_line(link):

	soup = oddsportal_soup(link + "#over-under")

	table = soup.select(".table-header-light")

	list_of_tables = []

	for line in table: 
		
		if str(line.select("span")[2].text):		# Added this parent "if" to prevent empty tables from raising "ValueError: could not convert string to float" on the code below.

			handicap = float(line.select("strong > a")[0].text[12:])		# Using str() to remove unicode 'u
			
			try:
				o_odds = float(line.select("span")[2].text)
			
			except ValueError:
				print(line.select("span")[2].text)
				o_odds = round(eval(line.select("span")[2].text.encode('utf-8')), 2 + 1)		# Sometimes UK odds appear "7/100" instead of floats "1.07". Eval() makes the division and "import __future__" exectutes all divisions as float divisions.

			try:
				u_odds = float(line.select("span")[1].text)

			except ValueError:
				u_odds = round(eval(line.select("span")[1].text.encode('utf-8')), 2 + 1)

			if 1.5 < o_odds < 3:		# Keep only the possibly usefull handicaps

				list_of_tables.append((handicap, o_odds, u_odds))


	l_o_t = list_of_tables
	
	line_odds = ()

	for i in range(len(l_o_t)):

		if l_o_t[i][1] >= 1.8:

			diff = l_o_t[i][1] - l_o_t[i][2]

			if abs(diff) < 0.25:

				line_odds = (l_o_t[i][0], l_o_t[i][1])

			elif l_o_t[i][0] - 0.25 != l_o_t[i-1][0] :		# if e.g. Over +3.25 does not exist, create it by averaging the odds of +3.5 and +3.

				line_odds = (l_o_t[i][0] - 0.25, (l_o_t[i][1] + l_o_t[i-1][1])/2)

			else:

				line_odds = (l_o_t[i][0], l_o_t[i][1])

			break
		


	return line_odds

# ------------------------------------------ START --------------------------------------------

# open spreadsheet

wb = load_workbook("data.xlsx")

wb.save("data_bak.xlsx")

ws = wb.active

last_row = ws.max_row

last_filled_row = last_row 								# In case xlsx is already full, prevents NameError (not defined) in the following for loop.

for i in range(last_row - 40, last_row):				# Gives last filled row of xlsx
	if ws.cell(row = i+1, column = 7).value == None:
		last_filled_row = i
		break

old_date = ""

for i in range(last_filled_row, last_row):

	new_date = 	str(ws.cell(row=i + 1, column=1).value)

	game = ws.cell(row=i + 1, column=4).value + " " + ws.cell(row=i + 1, column=5).value
	
	if datetime.datetime.strptime(new_date, "%Y%m%d").date() >= datetime.date.today():		# Don't check games with today's date, they are not played yet
		print('Done')
		exit()


	print("\n" + game)

	if old_date != new_date:
			
		soup = games_page_for_date(new_date)

		links_dic = links_dict(soup)

	game_link = find_game_link_in_page(links_dic, game)

	if not game_link:										# if game_link is not found (return None), do nothing and move to the next game in excel.
		
		ws.cell(row= i + 1, column= 7, value= "check").font = Font(color='0000ff')		# Blue text
		ws.cell(row= i + 1, column= 8, value= "date").font = Font(color='0000ff')

		continue

	game_score = get_score(soup, game_link)

	if game_score:

		line_odds = OU_line(game_link)

		if len(line_odds) < 2:					# Sometimes site does not load correclty and returns empty tuple. If so, re-load it.
			line_odds = OU_line(game_link)

		row_content = game_score + line_odds		# Merged what to write in the .xlsx row, on a single tuple.

		print(row_content)

		for j in range(4):

			ws.cell(row= i + 1, column= 7 + j, value= row_content[j])

		print("Written.")

	elif game_score == ():									# If score empty, game not played yet. Move on.
		
		print("Not played yet.")
		pass

	else:													# In none of the above, the game is postponed. Mark for manual removal.
		ws.cell(row= i + 1, column= 7, value= "postp.")
		ws.cell(row= i + 1, column= 8, value= "REMOVE")

		for row_cell in ws[i+1]:
			row_cell.font = Font(strike=True, italic=True, color='ff0000')		# Strikesthrugh, italisizes and colors RED any posponed game.

	wb.save("data.xlsx")
	
	old_date = new_date

print ("\n" + "Done! Match results up to date." + "\n")