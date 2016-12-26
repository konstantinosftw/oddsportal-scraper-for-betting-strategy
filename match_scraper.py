import urllib2
from bs4 import BeautifulSoup
import datetime
import pytz										# timezone library for UTC conversion
from openpyxl import Workbook, load_workbook
from os import path
import re


# ------------------------------------------ FUNCTIONS --------------------------------------------

# Returns the date, once it converts it from the header text. (e.g. today, tomorrow, 3rd Dec. etc)
def grab_date_from_header(header):

	header_text = header.text.encode('utf-8')		# .encode('utf-8') for unicode removal

	date_text = header_text.split()[-1]


	if date_text == "today":
			
		date = datetime.datetime.now(pytz.utc).date()		# pytz.utc for UTC conversion and date() to remove time parameters

	elif date_text == "tomorrow":
			
		date = (datetime.datetime.now(pytz.utc).date() + datetime.timedelta(days=1))		# adding timedelta(x) offsets time for given x

	else:
			
		header_date = header_text.split(', ', 1)[1]		# keeps only date part of header, which is after the first ", "

		purged_header_date = re.sub(r'\w{2}\,', '', header_date)		# removes letter suffix form day number - "2nd" becomes "2" etc

		date = datetime.datetime.strptime(purged_header_date, "%B %d %Y").date()
	
	return date


# Returns a list of lists that include the games for a given date [date, time, league, team 1, team 2, BTTS poins]
def football_match_rows(header, date):

	all_data = header.findNextSiblings()[0].select(".main-row")

	list_of_rows = []

	for x in all_data:

		temp_list = []

		temp_list.append(int(date.strftime('%Y%m%d')))								# strftime() formats the date as YYYYMMDD
		temp_list.append(x.select(".COL-1")[0].text.encode('utf-8'))				# .encode('utf-8') removes unicode e.g. [u'19:30'] - NEVER use str() for that purpose
		temp_list.append(x.select(".COL-2")[0].text.encode('utf-8').strip())		#strip() removes empty space from beginning and end of string
		temp_list.append(x.select(".COL-3")[0].text.encode('utf-8').strip())
		temp_list.append(x.select(".COL-5")[0].text.encode('utf-8').strip())
		temp_list.append(float(x.select(".COL-10")[0].text))

		if temp_list[5] >= 20:
			list_of_rows.append(temp_list)

	return list_of_rows


# Writes each list as a row in the .xlsx file
def fill_xlsx(list_of_rows):

	wb = load_workbook("data.xlsx")

	ws = wb.active

	start_from_row = ws.max_row + 1
	
	print('\nCopying...')

	for i in range(len(list_of_rows)):
		
		print(" ".join(map(str, list_of_rows[i])))		# map items to str() because can't join() floats

		for j in range(6):
			
			ws.cell(row=start_from_row + i, column=j+1, value=list_of_rows[i][j]) # j-th item from i-th list within the list_of_rows

	wb.save("data.xlsx")



# ------------------------------------------ START --------------------------------------------

print('Grabbing HTML page...')

url = urllib2.urlopen("http://www.over25tips.com/both-teams-to-score-tips")

content = url.read()

soup = BeautifulSoup(content,  "lxml")


if not path.exists("data.xlsx"):		# create file if not exists
	
	print('Creating Spreadsheet...')

	wb = Workbook()
	wb.save("data.xlsx")


print('Opening Spreadsheet...')

wb = load_workbook("data.xlsx")
ws = wb.active

last_date_cell = str(ws.cell(row=ws.max_row, column=1).value)		# get last date value from xlsx


if last_date_cell:

	last_date = datetime.datetime.strptime(last_date_cell, "%Y%m%d").date()		# the .date() suffix removes the "HH:MM:SS" parameters so that it can compare to other dates

else:

	last_date = datetime.datetime.strptime("19991212", "%Y%m%d").date()


header = soup.find_all("h3", "title-main")		# grabs the headers that include the dates

for i in header:

	header_date = grab_date_from_header(i)

	if header_date > last_date:

		list_of_rows = football_match_rows(i, header_date)

		fill_xlsx(list_of_rows)


# END